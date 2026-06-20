"""Guardrails loader — parse the Excel into a structured, enforceable ruleset.

Guardrails are a first-class layer (kickoff), not an afterthought. This module
reads two things from ``docs/Dream_Arabia_Persona_Training_Matrix_2.xlsx``:

1. Each persona tab's "6. AI BEHAVIOUR RULES & GUARDRAILS" table:
   Rule Type | Rule | Example / Trigger | Required Response Pattern | Severity
2. The cross-persona "Guardrails Matrix" tab:
   Domain | Risk | What Dream MUST NOT do | What Dream MUST do | Persona impact

It emits a normalised ``Ruleset`` of ``GuardrailRule`` records and can snapshot
them to ``config/guardrails.generated.json`` (committed) so the agent can run
without re-reading the workbook.

Severity:
- Persona behaviour rules carry explicit HARD / SOFT.
- Matrix domains carry HIGH / MEDIUM / LOW risk, mapped HIGH->HARD, else SOFT.
"""
from __future__ import annotations

import json
from dataclasses import asdict, dataclass, field
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = REPO_ROOT / "docs" / "Dream_Arabia_Persona_Training_Matrix_2.xlsx"
GENERATED_JSON = Path(__file__).resolve().parent / "guardrails.generated.json"

PERSONA_SHEETS = {
    "P1 — Investor": "P1",
    "P2 — Entrepreneur": "P2",
    "P3 — Tourist": "P3",
    "P4 — Foreign Talent": "P4",
    "P5 — Student": "P5",
    "P6 — Citizen Resident": "P6",
}
MATRIX_SHEET = "Guardrails Matrix"
_VALID_SEVERITY = {"HARD", "SOFT"}
_VALID_RISK = {"HIGH", "MEDIUM", "LOW"}


@dataclass(frozen=True)
class GuardrailRule:
    source: str            # "persona" | "matrix"
    persona: str | None    # persona id for persona rules; None for cross-persona
    personas: tuple[str, ...]  # personas the rule applies to
    domain: str | None     # matrix domain (None for persona rules)
    rule_type: str         # e.g. "NEVER confirm", or matrix "MUST NOT"/"MUST"
    rule: str              # the rule text / what to do
    trigger: str           # example/trigger that activates the rule
    response_pattern: str  # required response pattern
    severity: str          # "HARD" | "SOFT"
    risk: str | None       # matrix risk level (None for persona rules)

    def to_dict(self) -> dict:
        d = asdict(self)
        d["personas"] = list(self.personas)
        return d


@dataclass
class Ruleset:
    rules: list[GuardrailRule] = field(default_factory=list)

    def for_persona(self, persona_id: str) -> list[GuardrailRule]:
        """Rules that apply to a persona: its own tab + matrix rules tagging it."""
        return [r for r in self.rules if persona_id in r.personas]

    def hard(self, persona_id: str | None = None) -> list[GuardrailRule]:
        rules = self.for_persona(persona_id) if persona_id else self.rules
        return [r for r in rules if r.severity == "HARD"]

    def to_dict(self) -> dict:
        return {"rules": [r.to_dict() for r in self.rules]}

    @classmethod
    def from_dict(cls, data: dict) -> "Ruleset":
        rules = []
        for d in data.get("rules", []):
            d = dict(d)
            d["personas"] = tuple(d.get("personas", ()))
            rules.append(GuardrailRule(**d))
        return cls(rules=rules)


def _clean(v) -> str:
    return "" if v is None else str(v).strip()


def _parse_persona_impact(text: str) -> tuple[str, ...]:
    """Map a matrix 'Persona impact' cell to persona ids.

    Handles 'All personas', 'P1, P2, P4', 'P3 (visa)', 'P2, P4, P6 (employer-side)'.
    """
    text = text or ""
    if "all persona" in text.lower():
        return ("P1", "P2", "P3", "P4", "P5", "P6")
    found = []
    for token in ("P1", "P2", "P3", "P4", "P5", "P6"):
        if token in text and token not in found:
            found.append(token)
    return tuple(found)


def _find_header_row(ws, must_contain: set[str]) -> int | None:
    """Find the 1-based row whose lowercased cells contain all required labels."""
    want = {w.lower() for w in must_contain}
    for r in range(1, ws.max_row + 1):
        cells = {_clean(ws.cell(r, c).value).lower() for c in range(1, ws.max_column + 1)}
        if want <= cells:
            return r
    return None


def _column_index(ws, header_row: int, label: str) -> int | None:
    label = label.lower()
    for c in range(1, ws.max_column + 1):
        if _clean(ws.cell(header_row, c).value).lower() == label:
            return c
    return None


def _parse_persona_sheet(ws, persona_id: str) -> list[GuardrailRule]:
    header = _find_header_row(ws, {"rule type", "severity"})
    if header is None:
        return []
    cols = {
        name: _column_index(ws, header, name)
        for name in ("rule type", "rule", "example / trigger", "required response pattern", "severity")
    }
    sev_col = cols["severity"]
    rules: list[GuardrailRule] = []
    for r in range(header + 1, ws.max_row + 1):
        severity = _clean(ws.cell(r, sev_col).value).upper() if sev_col else ""
        if severity not in _VALID_SEVERITY:
            continue  # not a rule row (e.g. the Sample Conversation section)
        rules.append(GuardrailRule(
            source="persona",
            persona=persona_id,
            personas=(persona_id,),
            domain=None,
            rule_type=_clean(ws.cell(r, cols["rule type"]).value),
            rule=_clean(ws.cell(r, cols["rule"]).value),
            trigger=_clean(ws.cell(r, cols["example / trigger"]).value),
            response_pattern=_clean(ws.cell(r, cols["required response pattern"]).value),
            severity=severity,
            risk=None,
        ))
    return rules


def _parse_matrix_sheet(ws) -> list[GuardrailRule]:
    header = _find_header_row(ws, {"domain", "risk"})
    if header is None:
        return []
    cols = {
        "domain": _column_index(ws, header, "domain"),
        "risk": _column_index(ws, header, "risk"),
        "must_not": _column_index(ws, header, "what dream must not do"),
        "must": _column_index(ws, header, "what dream must do"),
        "impact": _column_index(ws, header, "persona impact"),
    }
    risk_col = cols["risk"]
    rules: list[GuardrailRule] = []
    for r in range(header + 1, ws.max_row + 1):
        risk = _clean(ws.cell(r, risk_col).value).upper() if risk_col else ""
        if risk not in _VALID_RISK:
            continue
        impact = _clean(ws.cell(r, cols["impact"]).value)
        personas = _parse_persona_impact(impact)
        rules.append(GuardrailRule(
            source="matrix",
            persona=None,
            personas=personas,
            domain=_clean(ws.cell(r, cols["domain"]).value),
            rule_type="MUST NOT",
            rule=_clean(ws.cell(r, cols["must_not"]).value),
            trigger=impact,
            response_pattern=_clean(ws.cell(r, cols["must"]).value),
            severity="HARD" if risk == "HIGH" else "SOFT",
            risk=risk,
        ))
    return rules


def load_ruleset(xlsx_path: str | Path = DEFAULT_XLSX) -> Ruleset:
    """Parse the workbook into a Ruleset (persona rules + matrix rules)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    rules: list[GuardrailRule] = []
    for sheet_name, persona_id in PERSONA_SHEETS.items():
        if sheet_name in wb.sheetnames:
            rules.extend(_parse_persona_sheet(wb[sheet_name], persona_id))
    if MATRIX_SHEET in wb.sheetnames:
        rules.extend(_parse_matrix_sheet(wb[MATRIX_SHEET]))
    wb.close()
    return Ruleset(rules=rules)


def load_generated(path: str | Path = GENERATED_JSON) -> Ruleset:
    """Load the committed snapshot (no workbook needed)."""
    return Ruleset.from_dict(json.loads(Path(path).read_text(encoding="utf-8")))


def write_generated(xlsx_path: str | Path = DEFAULT_XLSX, out: str | Path = GENERATED_JSON) -> Ruleset:
    """Parse the workbook and write the committed JSON snapshot."""
    ruleset = load_ruleset(xlsx_path)
    Path(out).write_text(json.dumps(ruleset.to_dict(), indent=2, ensure_ascii=False), encoding="utf-8")
    return ruleset


if __name__ == "__main__":
    rs = write_generated()
    by_sev = {s: len([r for r in rs.rules if r.severity == s]) for s in ("HARD", "SOFT")}
    print(f"Wrote {len(rs.rules)} rules to {GENERATED_JSON}  ({by_sev})")
