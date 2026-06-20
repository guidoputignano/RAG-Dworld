"""Load each persona tab's "Sample Conversation Flow" from the Excel.

These flows are the Definition-of-Done regression target: replay the user turns
against the agent and assert each answer cites sources (with year for statistics),
ends with the persona's handoff URL + a named action, and respects every HARD
guardrail for that persona.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl

from .guardrails_loader import DEFAULT_XLSX, PERSONA_SHEETS


@dataclass
class Turn:
    role: str   # "user" | "assistant"
    text: str


def _clean(v) -> str:
    return "" if v is None else str(v).strip()


def _parse_sheet(ws) -> list[Turn]:
    turns: list[Turn] = []
    in_section = False
    for r in range(1, ws.max_row + 1):
        a = _clean(ws.cell(r, 1).value)
        b = _clean(ws.cell(r, 2).value)
        if a.upper().startswith("7.") and "SAMPLE CONVERSATION" in a.upper():
            in_section = True
            continue
        if not in_section:
            continue
        role = a.lower()
        if role == "user" and b:
            turns.append(Turn("user", b))
        elif role == "dream" and b:
            turns.append(Turn("assistant", b))
    return turns


def load_sample_flows(xlsx_path: str | Path = DEFAULT_XLSX) -> dict[str, list[Turn]]:
    """persona_id -> ordered list of conversation turns from its tab."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    flows: dict[str, list[Turn]] = {}
    for sheet_name, persona_id in PERSONA_SHEETS.items():
        if sheet_name in wb.sheetnames:
            flows[persona_id] = _parse_sheet(wb[sheet_name])
    wb.close()
    return flows


def user_turns(xlsx_path: str | Path = DEFAULT_XLSX) -> dict[str, list[str]]:
    """persona_id -> the user messages only (what we replay against the agent)."""
    return {
        pid: [t.text for t in turns if t.role == "user"]
        for pid, turns in load_sample_flows(xlsx_path).items()
    }
